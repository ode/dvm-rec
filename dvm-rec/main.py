# Turn off bytecode generation
import sys
sys.dont_write_bytecode = True

# Django specific settings
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
import django
django.setup()
from db.models import *
from openpyxl import load_workbook

def main():
    wb = load_workbook(filename = 'Pilani21.xlsx')
    sheet = wb.active
    cur_course, cur_sec = Course(), Section()
    for i in range(3, sheet.max_row+1):
        print('{i}/{sheet.max_row}', end='\r')
        if (com_cod := sheet['A'+str(i)].value):
            cur_course = Course.objects.create(
                com_cod = com_cod,
                course_number = sheet['B'+str(i)].value,
                course_title = sheet['C'+str(i)].value,
                course_lec_units = int(x) if (x:=sheet['D'+str(i)].value) != '-' else 0,
                course_prac_units = int(x) if (x:= sheet['E'+str(i)].value) != '-' else 0,
                course_units = int(sheet['F'+str(i)].value),
            )
        else:
            if (sec := sheet.cell(row = i, column = 7).value):
                cur_sec = Section(
                    course = cur_course,
                    section_code = sec,
                    room = sheet['I'+str(i)].value,
                    compre_date = sheet['K'+str(i)].value,
                    days_and_hours = sheet['J'+str(i)].value
                )
                cur_sec.save()
                ins = Instructor(name=sheet['H'+str(i)].value)
                ins.save()
                cur_sec.instructors.add(ins)
            else: 
                ins = Instructor(name=sheet['H'+str(i)].value)
                ins.save()
                cur_sec.instructors.add(ins)

if __name__ == "__main__":
    main()